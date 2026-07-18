import csv
import io
import time
from typing import List, Dict, Any, Optional
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors

from services.dashboard_service import dashboard_service
from repositories.transaction_repository import transaction_repository
from repositories.analysis_repository import analysis_repository
from repositories.audit_repository import audit_repository
from services.metrics_service import metrics_service

class ReportService:
    def generate_csv_report(self, uid: str, company_id: Optional[str] = None) -> str:
        """Generates standard CSV output of all completed transactions."""
        start_time = time.time()
        txs = transaction_repository.list_transactions(uid, company_id)
        completed = [t for t in txs if t.get("status") == "completed"]
        
        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerow(["ID", "Date", "Merchant", "Category", "Amount", "Type", "Status", "Risk", "Notes"])
        
        for t in completed:
            writer.writerow([
                t.get("id"),
                t.get("date"),
                t.get("merchant"),
                t.get("category"),
                t.get("amount"),
                t.get("type"),
                t.get("status"),
                t.get("payment_risk", "low"),
                t.get("notes", "")
            ])
            
        metrics_service.record("firestore_duration", time.time() - start_time)
        return output.getvalue()

    def generate_excel_report(self, uid: str, company_id: Optional[str] = None) -> bytes:
        """Generates a beautifully branded and styled Excel report of transaction ledger and KPIs."""
        start_time = time.time()
        overview = dashboard_service.get_overview(uid, company_id)
        txs = transaction_repository.list_transactions(uid, company_id)
        completed = [t for t in txs if t.get("status") == "completed"]

        wb = Workbook()
        ws = wb.active
        ws.title = "Executive Summary"
        ws.views.sheetView[0].showGridLines = True

        # Custom Palette Fill Fills
        title_fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid") # Dark Blue
        kpi_label_fill = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")
        header_fill = PatternFill(start_color="475569", end_color="475569", fill_type="solid") # Slate

        # Custom Fonts
        title_font = Font(name="Arial", size=16, bold=True, color="FFFFFF")
        header_font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
        kpi_val_font = Font(name="Arial", size=12, bold=True, color="0F172A")
        kpi_lbl_font = Font(name="Arial", size=9, bold=True, color="475569")
        bold_font = Font(name="Arial", size=10, bold=True)
        regular_font = Font(name="Arial", size=10)

        # Borders
        thin_side = Side(border_style="thin", color="CBD5E1")
        thin_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

        # Title Block
        ws.merge_cells("A1:I2")
        title_cell = ws["A1"]
        title_cell.value = "REVENUE HUB - AI CFO REPORT"
        title_cell.font = title_font
        title_cell.fill = title_fill
        title_cell.alignment = Alignment(horizontal="center", vertical="center")

        ws.row_dimensions[1].height = 20
        ws.row_dimensions[2].height = 20
        
        # Spacer
        ws.append([])

        # KPI Block Labels
        ws.row_dimensions[4].height = 18
        ws.row_dimensions[5].height = 22
        
        kpis = [
            ("Total Revenue", f"₹{overview.get('totalRevenue', 0.0):,.2f}"),
            ("Total Expenses", f"₹{overview.get('totalExpenses', 0.0):,.2f}"),
            ("Net profit", f"₹{overview.get('netProfit', 0.0):,.2f}"),
            ("Business Health", f"{overview.get('healthScore', 94)}% ({overview.get('healthLabel', 'Stable')})")
        ]
        
        # Write KPI Block
        for col_idx, (lbl, val) in enumerate(kpis, start=1):
            cell_lbl = ws.cell(row=4, column=col_idx, value=lbl.upper())
            cell_lbl.font = kpi_lbl_font
            cell_lbl.fill = kpi_label_fill
            cell_lbl.alignment = Alignment(horizontal="center", vertical="center")
            cell_lbl.border = thin_border
            
            cell_val = ws.cell(row=5, column=col_idx, value=val)
            cell_val.font = kpi_val_font
            cell_val.alignment = Alignment(horizontal="center", vertical="center")
            cell_val.border = thin_border

        ws.append([])
        ws.append([])

        # Table Header
        headers = ["ID", "Date", "Merchant", "Category", "Amount", "Type", "Status", "Risk", "Notes"]
        ws.append(headers)
        ws.row_dimensions[8].height = 24
        
        for col_idx in range(1, 10):
            cell = ws.cell(row=8, column=col_idx)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border

        # Data Rows
        row_num = 9
        for t in completed:
            row_data = [
                t.get("id"),
                t.get("date"),
                t.get("merchant"),
                t.get("category"),
                t.get("amount"),
                t.get("type").upper(),
                t.get("status").upper(),
                t.get("payment_risk", "low").upper(),
                t.get("notes", "")
            ]
            ws.append(row_data)
            
            # Formatting
            ws.cell(row=row_num, column=5).number_format = '"₹"#,##0.00'
            for col_idx in range(1, 10):
                cell = ws.cell(row=row_num, column=col_idx)
                cell.font = regular_font
                cell.border = thin_border
                if col_idx in [1, 2, 6, 7, 8]:
                    cell.alignment = Alignment(horizontal="center")
            row_num += 1

        # Auto-fit columns
        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = openpyxl.utils.get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

        output = io.BytesIO()
        wb.save(output)
        
        metrics_service.record("firestore_duration", time.time() - start_time)
        return output.getvalue()

    def generate_pdf_report(self, uid: str, company_id: Optional[str] = None) -> bytes:
        """Generates a beautifully designed, highly branded PDF Executive CFO Report using ReportLab."""
        start_time = time.time()
        overview = dashboard_service.get_overview(uid, company_id)
        latest_analysis = analysis_repository.get_latest_analysis(uid, company_id) or {}
        analysis_data = latest_analysis.get("analysis", {})

        output = io.BytesIO()
        doc = SimpleDocTemplate(
            output,
            pagesize=letter,
            rightMargin=36,
            leftMargin=36,
            topMargin=36,
            bottomMargin=36
        )

        styles = getSampleStyleSheet()
        
        # Custom Typography Styles
        title_style = ParagraphStyle(
            'PDFTitle',
            parent=styles['Normal'],
            fontName='Helvetica-Bold',
            fontSize=22,
            leading=26,
            textColor=colors.HexColor('#1e3a8a'),
            spaceAfter=6
        )
        
        subtitle_style = ParagraphStyle(
            'PDFSubTitle',
            parent=styles['Normal'],
            fontName='Helvetica-Bold',
            fontSize=10,
            leading=12,
            textColor=colors.HexColor('#0d9488'),
            spaceAfter=15,
            textTransform='uppercase'
        )

        h1_style = ParagraphStyle(
            'PDFH1',
            parent=styles['Heading2'],
            fontName='Helvetica-Bold',
            fontSize=12,
            leading=14,
            textColor=colors.HexColor('#0f172a'),
            spaceBefore=14,
            spaceAfter=8,
            borderPadding=(0, 0, 2, 0),
            borderColor=colors.HexColor('#cbd5e1'),
            borderWidth=0.5
        )

        body_style = ParagraphStyle(
            'PDFBody',
            parent=styles['Normal'],
            fontName='Helvetica',
            fontSize=9.5,
            leading=13,
            textColor=colors.HexColor('#334155'),
            spaceAfter=8
        )

        kpi_val_style = ParagraphStyle(
            'KpiVal',
            parent=styles['Normal'],
            fontName='Helvetica-Bold',
            fontSize=14,
            leading=16,
            textColor=colors.HexColor('#0f172a'),
            alignment=1 # Center
        )

        kpi_lbl_style = ParagraphStyle(
            'KpiLbl',
            parent=styles['Normal'],
            fontName='Helvetica-Bold',
            fontSize=8,
            leading=10,
            textColor=colors.HexColor('#64748b'),
            alignment=1, # Center
            spaceAfter=4
        )

        story = []

        # Branded Header
        story.append(Paragraph("REVENUE HUB", title_style))
        story.append(Paragraph("AI CFO EXECUTIVE DISPATCH & ANALYSIS REPORT", subtitle_style))
        story.append(Spacer(1, 10))

        # KPI Block Grid
        kpi_data = [
            [
                Paragraph("TOTAL REVENUE", kpi_lbl_style),
                Paragraph("TOTAL EXPENSES", kpi_lbl_style),
                Paragraph("NET PROFIT", kpi_lbl_style),
                Paragraph("HEALTH SCORE", kpi_lbl_style)
            ],
            [
                Paragraph(f"₹{overview.get('totalRevenue', 0.0):,.0f}", kpi_val_style),
                Paragraph(f"₹{overview.get('totalExpenses', 0.0):,.0f}", kpi_val_style),
                Paragraph(f"₹{overview.get('netProfit', 0.0):,.0f}", kpi_val_style),
                Paragraph(f"{overview.get('healthScore', 94)}% ({overview.get('healthLabel', 'Stable')})", kpi_val_style)
            ]
        ]
        
        kpi_table = Table(kpi_data, colWidths=[135, 135, 135, 135])
        kpi_table.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,-1), colors.HexColor('#f8fafc')),
            ('BOX', (0,0), (-1,-1), 1, colors.HexColor('#e2e8f0')),
            ('INNERGRID', (0,0), (-1,-1), 0.5, colors.HexColor('#e2e8f0')),
            ('TOPPADDING', (0,0), (-1,-1), 12),
            ('BOTTOMPADDING', (0,0), (-1,-1), 12),
            ('ALIGN', (0,0), (-1,-1), 'CENTER'),
            ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ]))
        story.append(kpi_table)
        story.append(Spacer(1, 15))

        # Executive Summary
        story.append(Paragraph("Executive Summary Insights", h1_style))
        summary_text = analysis_data.get("summary", "No AI summary compiled yet. Upload financial ledger statements to compile forecasts.")
        story.append(Paragraph(summary_text, body_style))

        # AI CFO Recommendations
        story.append(Paragraph("Strategic CFO Action Items", h1_style))
        recs = analysis_data.get("recommendations", [])
        if not recs:
            story.append(Paragraph("No recommendation items currently compiled.", body_style))
        else:
            for idx, r in enumerate(recs, start=1):
                p_text = f"<b>{idx}. [{r.get('priority', 'Medium')}]</b> {r.get('action')}"
                story.append(Paragraph(p_text, body_style))

        # Audited Risks
        story.append(Paragraph("Audited Operational Risks", h1_style))
        risks_list = analysis_data.get("risks", [])
        if not risks_list:
            story.append(Paragraph("No significant risks registered.", body_style))
        else:
            # Build Risk Table
            table_data = [[
                Paragraph("<b>Risk</b>", body_style),
                Paragraph("<b>Severity</b>", body_style),
                Paragraph("<b>Impact</b>", body_style),
                Paragraph("<b>Recommendation</b>", body_style)
            ]]
            for r in risks_list:
                table_data.append([
                    Paragraph(r.get("risk", ""), body_style),
                    Paragraph(r.get("severity", ""), body_style),
                    Paragraph(r.get("financialImpact", ""), body_style),
                    Paragraph(r.get("recommendation", ""), body_style)
                ])
            
            risk_table = Table(table_data, colWidths=[130, 60, 110, 240])
            risk_table.setStyle(TableStyle([
                ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#cbd5e1')),
                ('BOX', (0,0), (-1,-1), 0.5, colors.HexColor('#94a3b8')),
                ('INNERGRID', (0,0), (-1,-1), 0.5, colors.HexColor('#cbd5e1')),
                ('TOPPADDING', (0,0), (-1,-1), 6),
                ('BOTTOMPADDING', (0,0), (-1,-1), 6),
                ('VALIGN', (0,0), (-1,-1), 'TOP'),
            ]))
            story.append(risk_table)

        # Forecast Periods
        story.append(Paragraph("Future Month Forecasts", h1_style))
        forecasts = analysis_data.get("forecast", {})
        if not forecasts:
            story.append(Paragraph("No forecast projections available.", body_style))
        else:
            fc_data = [["Month Period", "Projected Revenue Target"]]
            for k, v in forecasts.items():
                try:
                    fc_data.append([k, f"₹{float(v):,.2f}"])
                except:
                    fc_data.append([k, str(v)])
                    
            fc_table = Table(fc_data, colWidths=[200, 200])
            fc_table.setStyle(TableStyle([
                ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#f1f5f9')),
                ('BOX', (0,0), (-1,-1), 0.5, colors.HexColor('#e2e8f0')),
                ('INNERGRID', (0,0), (-1,-1), 0.5, colors.HexColor('#e2e8f0')),
                ('TOPPADDING', (0,0), (-1,-1), 5),
                ('BOTTOMPADDING', (0,0), (-1,-1), 5),
            ]))
            story.append(fc_table)

        # Build document
        doc.build(story)
        pdf_bytes = output.getvalue()
        
        metrics_service.record("firestore_duration", time.time() - start_time)
        return pdf_bytes

report_service = ReportService()
